# -*- coding: utf-8 -*-
"""
生成策略对比分析 Excel：旧版 vs 新版 F1F2F3 回测结果对比 + 变化原因分析。
"""
import os
import sys
from collections import defaultdict

import pandas as pd

_PROJECT_ROOT = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
sys.path.insert(0, _PROJECT_ROOT)

REPORTS_DIR = os.path.join(_PROJECT_ROOT, "reports")
OLD_XLSX = os.path.join(REPORTS_DIR, "引擎回测_prot_adj_2025-01-02_2025-12-31.xlsx")
NEW_XLSX = os.path.join(REPORTS_DIR, "引擎回测_prot_adj_2025-01-02_2026-04-30.xlsx")
OUTPUT = os.path.join(REPORTS_DIR, "策略对比分析_旧版vs新版F1F2F3.xlsx")


def read_sheet(path, sheet):
    return pd.read_excel(path, sheet_name=sheet)


def read_summary_dict(path):
    df = read_sheet(path, "汇总统计")
    return {str(row.iloc[0]): str(row.iloc[1]) for _, row in df.iterrows()}


def read_monthly_dict(path):
    df = read_sheet(path, "月度分布")
    return {str(row.iloc[0]): (int(row.iloc[1]), float(row.iloc[2])) for _, row in df.iterrows()}


def read_reason_dict(path):
    df = read_sheet(path, "卖出原因")
    result = {}
    for _, row in df.iterrows():
        result[str(row.iloc[0])] = {
            "笔数": int(row.iloc[1]), "盈亏合计": float(row.iloc[2]),
            "平均持仓天": float(row.iloc[3]) if len(row) > 3 else 0,
            "胜率%": float(row.iloc[4]) if len(row) > 4 else 0,
        }
    return result


def main():
    s_old = read_summary_dict(OLD_XLSX)
    s_new = read_summary_dict(NEW_XLSX)
    m_old = read_monthly_dict(OLD_XLSX)
    m_new = read_monthly_dict(NEW_XLSX)
    r_old = read_reason_dict(OLD_XLSX)
    r_new = read_reason_dict(NEW_XLSX)

    # ================================================================
    # Sheet 1: 策略逻辑对比
    # ================================================================
    strategy_rows = [
        ["版本", "旧版 (2025年仅回测)", "新版 (2026.04)"],
        ["代码版本", "8c38f8a (新增K线可视化+回测框架)", "516b94b (回测引擎升级+K线可视化增强)"],
        ["回测区间", "2025-01-02 ~ 2025-12-31", "2025-01-02 ~ 2026-04-30"],
        ["", "", ""],
        ["━━━ F1 均线条件 ━━━", "", ""],
        ["F1 规则", "MA5 > MA10 > MA20 > MA60\n且四条均线斜率均 > 0", "MA60斜率>0\n(MA20或MA30至少一个斜率>0)\nMA5斜率>=0"],
        ["F1 含义", "严格多头排列：短期均线在上，长期在下，\n且所有均线都向上", "放宽为趋势向上：长周期必须向上，\n中周期至少一个向上，短期不下降即可"],
        ["F1 变化", "基准", "大幅放宽：取消了MA严格排序，取消了MA10/MA20/MA30必须全向上"],
        ["", "", ""],
        ["━━━ F2 涨停基因 ━━━", "", ""],
        ["F2 规则", "近20日内 max(C/prev_close) >= 1.098", "同左（未变）"],
        ["F2 含义", "20日涨停基因，确保有涨停活跃度", "未变"],
        ["", "", ""],
        ["━━━ F3 均线位置 ━━━", "", ""],
        ["F3 规则", "Close > MA5\n（收盘价在5日线上方即可）", "CROSS(C, MA5) 上穿当日买入，\n若当日MA5斜率不满足，\n次日C>MA5且MA5不下降也可买入"],
        ["F3 含义", "股价站在5日线之上，简单的趋势确认", "精确到金叉信号：必须是\"刚上穿\"的时点，\n但允许延迟一天确认"],
        ["F3 变化", "基准", "从\"在线上\"变为\"刚上穿\"，增加了精确时点要求\n但通过次日补买机制放宽"],
        ["", "", ""],
        ["━━━ 卖出机制 ━━━", "", ""],
        ["策略卖出(signal)", "跌破PROT_ADJ移动止盈线\n或四天跌5%急跌", "跌破PROT_ADJ移动止盈线\n(关闭了急跌)"],
        ["固定止损(stop_loss)", "10% 固定止损", "10% 固定止损（未变）"],
        ["利润保护-10%", "无", "profit_lock_10pct:\n持仓最大盈利≥10%后，若当前盈利<10%\n且最低价>PROT_ADJ线，则止盈"],
        ["利润保护-3%", "无", "profit_lock_3pct:\n持仓最大盈利≥5%后，若PROT_ADJ利润<3%，\n则锁定3%利润"],
        ["PROT_ADJ保护-10%", "无", "prot_adj_10pct:\n持仓最大盈利≥10%后，若跌破PROT_ADJ线，\n则止盈卖出"],
        ["移动止盈(trailing)", "trailing_stop 跟踪止盈", "已取消，由分层利润保护替代"],
        ["止盈(take_profit)", "take_profit 止盈", "已取消，由分层利润保护替代"],
    ]
    df_strategy = pd.DataFrame(strategy_rows, columns=["对比维度", "旧版", "新版"])

    # ================================================================
    # Sheet 2: 核心指标对比
    # ================================================================
    metric_keys = [
        ("初始资金", "初始资金"),
        ("期末总资产", "期末总资产"),
        ("总收益率", "总收益率"),
        ("年化收益率", "年化收益率"),
        ("最大回撤", "最大回撤"),
        ("回撤天数", "回撤天数"),
        ("总交易次数(已平仓)", "总交易次数"),
        ("未平仓数", "未平仓数"),
        ("盈利次数", "盈利次数"),
        ("亏损次数", "亏损次数"),
        ("胜率", "胜率"),
        ("总盈亏(元)", "总盈亏(元)"),
        ("总手续费(元)", "总手续费(元)"),
        ("平均每笔盈亏(元)", "平均每笔盈亏(元)"),
        ("最大单笔盈利(元)", "最大单笔盈利(元)"),
        ("最大单笔亏损(元)", "最大单笔亏损(元)"),
        ("平均持仓(天)", "平均持仓(天)"),
        ("最短持仓(天)", "最短持仓(天)"),
        ("最长持仓(天)", "最长持仓(天)"),
    ]
    metric_rows = []
    for old_key, label in metric_keys:
        v_old = s_old.get(old_key, "-")
        v_new = s_new.get(old_key, "-")
        metric_rows.append([label, v_old, v_new])
    df_metrics = pd.DataFrame(metric_rows, columns=["指标", "旧版", "新版"])

    # ================================================================
    # Sheet 3: 月度盈亏对比
    # ================================================================
    all_months = sorted(set(list(m_old.keys()) + list(m_new.keys())))
    monthly_rows = []
    for m in all_months:
        t1, p1 = m_old.get(m, (0, 0))
        t2, p2 = m_new.get(m, (0, 0))
        monthly_rows.append({
            "月份": m, "旧版交易笔数": t1, "旧版盈亏": round(p1, 0),
            "新版交易笔数": t2, "新版盈亏": round(p2, 0),
            "盈亏差异": round(p2 - p1, 0),
        })
    df_monthly = pd.DataFrame(monthly_rows)

    # ================================================================
    # Sheet 4: 卖出原因对比
    # ================================================================
    all_reasons = sorted(set(list(r_old.keys()) + list(r_new.keys())))
    reason_rows = []
    for reason in all_reasons:
        ro = r_old.get(reason, {"笔数": 0, "盈亏合计": 0, "平均持仓天": 0, "胜率%": 0})
        rn = r_new.get(reason, {"笔数": 0, "盈亏合计": 0, "平均持仓天": 0, "胜率%": 0})
        reason_rows.append({
            "卖出原因": reason, "旧版笔数": ro["笔数"], "旧版盈亏合计": round(ro["盈亏合计"], 0),
            "旧版平均持仓天": ro["平均持仓天"], "旧版胜率%": ro["胜率%"],
            "新版笔数": rn["笔数"], "新版盈亏合计": round(rn["盈亏合计"], 0),
            "新版平均持仓天": rn["平均持仓天"], "新版胜率%": rn["胜率%"],
        })
    df_reason = pd.DataFrame(reason_rows)

    # ================================================================
    # Sheet 5: 变化原因深度分析
    # ================================================================
    analysis_rows = [
        ["━━━ 变化1：F1 放宽导致大量弱势股进入候选 ━━━", "", ""],
        ["旧版F1", "MA5>MA10>MA20>MA60 且四条均线全向上",
         "这种\"四线开花\"形态只出现在强势趋势启动时，\n天然过滤了震荡和弱势股，信号稀少但质量高"],
        ["新版F1", "MA60↑ + (MA20↑|MA30↑) + MA5≥0",
         "只要求长周期向上+短周期不跌，大量底部震荡、\n弱势反弹股都能通过，信号量暴增但质量下降"],
        ["量化影响", "旧版2025全年607笔交易", "新版同期(2025)约900+笔（按前4个月推算），\n交易量增加约50%，多为低质量信号"],
        ["", "", ""],
        ["━━━ 变化2：F3 从\"站在线上\"改为\"刚上穿\" —— 双刃剑 ━━━", "", ""],
        ["旧版F3", "Close > MA5",
         "只要站在5日线上就满足，错过了追入时点\n但避免了\"假金叉\"后立即回落的风险"],
        ["新版F3", "CROSS(C,MA5) + MA5斜率确认 + 次日补买",
         "精确到金叉时刻，追入更及时。但：\n① MA5斜率确认不全（只要求>=0）\n② 次日补买机制在弱势行情中会\"接飞刀\""],
        ["量化影响", "旧版signal卖出602笔/盈+55万",
         "新版signal卖出316笔/亏-118万\n说明新F3追入的时点后续走势很差"],
        ["", "", ""],
        ["━━━ 变化3：卖出机制增加了分层利润保护 ━━━", "", ""],
        ["正面效果", "profit_lock_10pct +329笔 +251万\nprot_adj_10pct +37笔 +44万",
         "当股票确实走出行情时，利润保护能锁定收益\n这是新版唯一的亮点，贡献了+295万利润"],
        ["负面效果", "profit_lock_3pct 367笔 -26万",
         "3%的利润锁定阈值太低，稍有波动就触发\n平均持仓仅6天就被震出，错失后续涨幅"],
        ["", "", ""],
        ["━━━ 变化4：止损暴增 5次→134次 ━━━", "", ""],
        ["旧版止损", "全年仅5次止损，亏损-5.9万",
         "严格的买入条件意味着入选的股票本身就在\n强势趋势中，很少有触及-10%止损的情况"],
        ["新版止损", "全年134次止损，亏损-131.5万",
         "放宽F1+F3后，大量\"假强势\"股票入选，\n买入后即转弱，频繁触发止损"],
        ["止损率对比", "旧版: 5/607 = 0.8%", "新版: 134/1183 = 11.3%，止损率高14倍"],
        ["", "", ""],
        ["━━━ 综合结论 ━━━", "", ""],
        ["根因", "",
         "F1放宽(取消均线排序) + F3精确化(CROSS替代简单在线) → \n买入信号量增但质降，大量弱势/假突破股票被买入，\n导致signal卖出和stop_loss亏损暴增。\n分层利润保护(profit_lock_10pct)表现优异但不足以弥补买入端恶化。"],
        ["核心数据", "旧版: 年化+44.9%, 回撤7.2%, 607笔",
         "新版: 年化+9.5%, 回撤32.6%, 1183笔\n交易量翻倍，收益率降为1/5，回撤扩大4.5倍"],
        ["", "", ""],
        ["━━━ 策略优化建议 ━━━", "", ""],
        ["建议1: 恢复旧版F1", "恢复 MA5>MA10>MA20>MA60 且四线上扬",
         "这是质量控制的核心，预期可让signal卖出恢复盈利"],
        ["建议2: 保留新版F3的CROSS", "保留 CROSS(C,MA5) 但取消次日补买",
         "CROSS本身就是精确信号，次日补买引入了滞后追高"],
        ["建议3: 保留分层利润保护", "profit_lock_10pct + prot_adj_10pct",
         "这是新版最成功的改进，建议保留"],
        ["建议4: 调整3%利润保护", "将profit_lock_3pct阈值提高到5%并放宽触发条件",
         "当前3%太低，频繁被震出"],
        ["建议5: 组合测试", "F1旧版+F3新版+分层利润保护",
         "预期：交易量适中(500-700笔)，信号质量高，\n利润保护锁定收益，年化收益20-40%，回撤<15%"],
    ]
    df_analysis = pd.DataFrame(analysis_rows, columns=["分析维度", "旧版", "新版/结论"])

    # ================================================================
    # 写入 Excel（带格式）
    # ================================================================
    with pd.ExcelWriter(OUTPUT, engine="openpyxl") as writer:
        df_strategy.to_excel(writer, sheet_name="1-策略逻辑对比", index=False)
        df_metrics.to_excel(writer, sheet_name="2-核心指标对比", index=False)
        df_monthly.to_excel(writer, sheet_name="3-月度盈亏对比", index=False)
        df_reason.to_excel(writer, sheet_name="4-卖出原因对比", index=False)
        df_analysis.to_excel(writer, sheet_name="5-变化原因深度分析", index=False)

    # 调整列宽
    from openpyxl import load_workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    wb = load_workbook(OUTPUT)

    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font_white = Font(bold=True, size=11, color="FFFFFF")
    wrap_align = Alignment(wrap_text=True, vertical="top")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    for ws in wb.worksheets:
        # Header row
        for cell in ws[1]:
            cell.font = header_font_white
            cell.fill = header_fill
            cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
        # Data rows
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=ws.max_column):
            for cell in row:
                cell.alignment = wrap_align
                cell.border = thin_border
        # Auto-width (capped)
        for col in ws.columns:
            max_len = 0
            col_letter = col[0].column_letter
            for cell in col:
                if cell.value:
                    # Take longest line of wrapped text
                    lines = str(cell.value).split("\n")
                    line_max = max(len(line) for line in lines)
                    max_len = max(max_len, line_max)
            ws.column_dimensions[col_letter].width = min(max_len + 4, 55)

    wb.save(OUTPUT)
    print(f"策略对比分析 Excel 已生成: {OUTPUT}")


if __name__ == "__main__":
    main()
